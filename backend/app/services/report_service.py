import io
import uuid
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.ho_so import HoSoViPham
from app.models.linh_vuc import LinhVucViPham
from app.models.thon import Thon

HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
BOLD = Font(bold=True)


async def build_bao_cao_ma_tran(
    db: AsyncSession,
    tu_ngay: datetime,
    den_ngay: datetime,
    thon_id: uuid.UUID | None = None,
) -> bytes:
    """Sinh file Excel ma tran: hang doc = Thon, cot ngang = Linh vuc VPHC.
    Moi o giao nhau gom 2 gia tri: Tong so vu + Tong so tien phat.
    """
    thon_stmt = select(Thon).where(Thon.trang_thai == "hoat_dong").order_by(Thon.ten_thon)
    if thon_id is not None:
        thon_stmt = select(Thon).where(Thon.id == thon_id)
    thon_list = (await db.execute(thon_stmt)).scalars().all()

    linh_vuc_list = (
        (await db.execute(select(LinhVucViPham).order_by(LinhVucViPham.thu_tu_hien_thi)))
        .scalars()
        .all()
    )

    agg_stmt = (
        select(
            HoSoViPham.thon_id,
            HoSoViPham.linh_vuc_id,
            func.count(HoSoViPham.id).label("so_vu"),
            func.coalesce(func.sum(HoSoViPham.so_tien_phat), 0).label("tong_tien"),
        )
        .where(HoSoViPham.ngay_lap >= tu_ngay, HoSoViPham.ngay_lap <= den_ngay)
        .group_by(HoSoViPham.thon_id, HoSoViPham.linh_vuc_id)
    )
    if thon_id is not None:
        agg_stmt = agg_stmt.where(HoSoViPham.thon_id == thon_id)

    rows = (await db.execute(agg_stmt)).all()
    agg_map: dict[tuple[uuid.UUID, uuid.UUID], tuple[int, float]] = {
        (r.thon_id, r.linh_vuc_id): (r.so_vu, float(r.tong_tien)) for r in rows
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Bao cao VPHC"

    ws.merge_cells("A1:A2")
    ws["A1"] = "Thon / Dia ban"
    ws["A1"].font = BOLD
    ws["A1"].alignment = Alignment(vertical="center", horizontal="center")

    col = 2
    for lv in linh_vuc_list:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        cell = ws.cell(row=1, column=col, value=lv.ten_linh_vuc)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.cell(row=2, column=col, value="So vu").font = BOLD
        ws.cell(row=2, column=col + 1, value="So tien phat (VND)").font = BOLD
        col += 2

    total_col = col
    ws.merge_cells(start_row=1, start_column=total_col, end_row=1, end_column=total_col + 1)
    ws.cell(row=1, column=total_col, value="Tong cong").font = BOLD
    ws.cell(row=2, column=total_col, value="So vu").font = BOLD
    ws.cell(row=2, column=total_col + 1, value="So tien phat (VND)").font = BOLD

    row_idx = 3
    grand_total_vu = 0
    grand_total_tien = 0.0
    col_totals: dict[int, tuple[int, float]] = {}

    for thon in thon_list:
        ws.cell(row=row_idx, column=1, value=thon.ten_thon)
        col = 2
        row_total_vu = 0
        row_total_tien = 0.0
        for lv in linh_vuc_list:
            so_vu, tong_tien = agg_map.get((thon.id, lv.id), (0, 0.0))
            ws.cell(row=row_idx, column=col, value=so_vu)
            ws.cell(row=row_idx, column=col + 1, value=tong_tien)
            prev_vu, prev_tien = col_totals.get(col, (0, 0.0))
            col_totals[col] = (prev_vu + so_vu, prev_tien + tong_tien)
            row_total_vu += so_vu
            row_total_tien += tong_tien
            col += 2

        ws.cell(row=row_idx, column=total_col, value=row_total_vu).font = BOLD
        ws.cell(row=row_idx, column=total_col + 1, value=row_total_tien).font = BOLD
        grand_total_vu += row_total_vu
        grand_total_tien += row_total_tien
        row_idx += 1

    ws.cell(row=row_idx, column=1, value="TONG CONG TOAN XA").font = BOLD
    col = 2
    for lv in linh_vuc_list:
        vu, tien = col_totals.get(col, (0, 0.0))
        ws.cell(row=row_idx, column=col, value=vu).font = BOLD
        ws.cell(row=row_idx, column=col + 1, value=tien).font = BOLD
        col += 2
    ws.cell(row=row_idx, column=total_col, value=grand_total_vu).font = BOLD
    ws.cell(row=row_idx, column=total_col + 1, value=grand_total_tien).font = BOLD

    ws.column_dimensions["A"].width = 28
    for c in range(2, total_col + 2):
        ws.column_dimensions[get_column_letter(c)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
